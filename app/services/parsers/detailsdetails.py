"""
DetailsDetails XLSX Parser

Parses DetailsDetails music sales reports (Digital Sales + Physical Sales sheets).
Returns raw parsed data for normalization.

DetailsDetails XLSX format:
- Two sheets: "Digital Sales" and "Physical Sales"
- Uses openpyxl for XLSX parsing
- Sales Period format: "2025-M06" (year-Mmonth)
- Royalty Rate as percentage string: "25.00%"
"""

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional


@dataclass
class DetailsDetailsRow:
    """Raw parsed row from DetailsDetails XLSX."""
    row_number: int
    sheet_name: str  # "Digital Sales" or "Physical Sales"
    contract_id: Optional[int]
    contract_name: Optional[str]
    artist: str
    title: str
    album_title: Optional[str]
    barcode: Optional[str]  # UPC (Barcode for digital, Identifier for physical)
    isrc: Optional[str]
    country: Optional[str]
    shop: Optional[str]  # Store name
    sales_period: Optional[str]  # "2025-M06"
    usage_type: Optional[str]  # "Track Stream", "Physical", etc.
    sales: int
    returns: int
    ppu: Decimal  # Price per unit
    amount: Decimal  # Gross amount
    royalty_rate: Optional[Decimal]  # Parsed from "25.00%" -> 0.25
    royalty_amount: Decimal
    physical_format: Optional[str]  # "LP", "CD", etc. (Physical Sales only)


@dataclass
class ParseError:
    """Represents a parsing error for a specific row."""
    row_number: int
    error: str
    raw_data: Optional[Dict] = None


@dataclass
class DetailsDetailsParseResult:
    """Result of parsing a DetailsDetails XLSX."""
    rows: List[DetailsDetailsRow] = field(default_factory=list)
    errors: List[ParseError] = field(default_factory=list)
    total_rows: int = 0


def _parse_decimal(value) -> Decimal:
    """Parse a decimal value from various types."""
    if value is None:
        return Decimal("0")

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return Decimal("0")
        # Remove currency symbols
        cleaned = cleaned.replace("€", "").replace("$", "").replace("£", "").strip()
        # Remove thousands separators (space or thin space)
        cleaned = cleaned.replace(" ", "").replace("\u202f", "")
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            raise ValueError(f"Cannot parse decimal: {value}")

    return Decimal(str(value))


def _parse_percentage(value) -> Optional[Decimal]:
    """
    Parse a percentage string like "25.00%" -> 0.25.
    Also handles numeric values (already as decimal or as percent number).
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        # If value is between 0 and 1, assume already a ratio
        if 0 <= value <= 1:
            return Decimal(str(value))
        # Otherwise assume it's a percentage number (25.0 -> 0.25)
        return Decimal(str(value)) / Decimal("100")

    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        if cleaned.endswith("%"):
            cleaned = cleaned[:-1].strip()
            try:
                return Decimal(cleaned) / Decimal("100")
            except InvalidOperation:
                return None
        try:
            val = Decimal(cleaned)
            if val > 1:
                return val / Decimal("100")
            return val
        except InvalidOperation:
            return None

    return None


def _parse_int(value) -> int:
    """Parse an integer value from various types."""
    if value is None:
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return 0
        try:
            return int(float(cleaned))
        except ValueError:
            raise ValueError(f"Cannot parse integer: {value}")
    return int(value)


def _safe_str(value) -> Optional[str]:
    """Safely convert a value to string, return None for empty/None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _find_header_indices(header_row: list, mappings: Dict[str, List[str]]) -> Dict[str, Optional[int]]:
    """Find column indices from header row using mappings."""
    headers_lower = [str(h).lower().strip() if h else "" for h in header_row]
    indices = {}

    for field_name, possible_names in mappings.items():
        indices[field_name] = None
        for name in possible_names:
            if name.lower() in headers_lower:
                indices[field_name] = headers_lower.index(name.lower())
                break

    return indices


# Column name mappings for Digital Sales sheet
DIGITAL_COLUMN_MAPPINGS = {
    "contract_id": ["contract id"],
    "contract_name": ["contract name"],
    "album_title": ["album title"],
    "barcode": ["barcode"],
    "isrc": ["isrc"],
    "artist": ["artist"],
    "title": ["title"],
    "usage_type": ["usage type"],
    "country": ["country"],
    "shop": ["shop"],
    "sales_period": ["sales period"],
    "sales": ["sales"],
    "returns": ["returns"],
    "ppu": ["ppu"],
    "amount": ["amount"],
    "royalty_rate": ["royalty rate"],
    "royalty_amount": ["royalty amount"],
}

# Column name mappings for Physical Sales sheet
PHYSICAL_COLUMN_MAPPINGS = {
    "contract_id": ["contract id"],
    "contract_name": ["contract name"],
    "identifier": ["identifier"],
    "artist": ["artist"],
    "title": ["title"],
    "format": ["format"],
    "type": ["type"],
    "country": ["country"],
    "sales": ["sales"],
    "returns": ["returns"],
    "ppu": ["ppu"],
    "amount": ["amount"],
    "royalty_rate": ["royalty rate"],
    "royalty_amount": ["royalty amount"],
}


class DetailsDetailsParser:
    """Parser for DetailsDetails XLSX files."""

    def parse(self, file_bytes: bytes) -> DetailsDetailsParseResult:
        """
        Parse DetailsDetails XLSX content.

        Args:
            file_bytes: XLSX file content as bytes

        Returns:
            DetailsDetailsParseResult with parsed rows and errors
        """
        import io

        from openpyxl import load_workbook

        result = DetailsDetailsParseResult()

        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            result.errors.append(ParseError(
                row_number=0,
                error=f"Cannot open XLSX file: {str(e)}",
            ))
            return result

        # Parse Digital Sales sheet
        self._parse_digital_sales(wb, result)

        # Parse Physical Sales sheet
        self._parse_physical_sales(wb, result)

        wb.close()
        return result

    def _parse_digital_sales(self, wb, result: DetailsDetailsParseResult) -> None:
        """Parse the Digital Sales sheet."""
        sheet_name = "Digital Sales"

        # Try to find the sheet (case-insensitive)
        ws = None
        for name in wb.sheetnames:
            if name.lower().strip() == sheet_name.lower():
                ws = wb[name]
                break

        if ws is None:
            # Try first sheet as fallback
            if len(wb.sheetnames) > 0:
                ws = wb[wb.sheetnames[0]]
                sheet_name = wb.sheetnames[0]
            else:
                result.errors.append(ParseError(
                    row_number=0,
                    error="No sheets found in XLSX file",
                ))
                return

        rows_iter = ws.iter_rows(values_only=True)

        # Get header row
        try:
            header_row = next(rows_iter)
        except StopIteration:
            result.errors.append(ParseError(
                row_number=0,
                error=f"Sheet '{sheet_name}' is empty",
            ))
            return

        header_list = list(header_row)
        indices = _find_header_indices(header_list, DIGITAL_COLUMN_MAPPINGS)

        # Validate required columns
        required = ["artist", "amount"]
        missing = [f for f in required if indices.get(f) is None]
        if missing:
            result.errors.append(ParseError(
                row_number=0,
                error=f"Missing required columns in '{sheet_name}': {missing}. Available: {[str(h) for h in header_list if h]}",
            ))
            return

        for row_number, row in enumerate(rows_iter, start=2):
            result.total_rows += 1
            row_list = list(row)

            # Skip empty rows
            if not row_list or all(cell is None or str(cell).strip() == "" for cell in row_list):
                continue

            try:
                artist = _safe_str(row_list[indices["artist"]] if indices["artist"] is not None else None)
                if not artist:
                    continue  # Skip rows without artist

                parsed_row = DetailsDetailsRow(
                    row_number=row_number,
                    sheet_name=sheet_name,
                    contract_id=_parse_int(row_list[indices["contract_id"]]) if indices["contract_id"] is not None else None,
                    contract_name=_safe_str(row_list[indices["contract_name"]] if indices["contract_name"] is not None else None),
                    artist=artist,
                    title=_safe_str(row_list[indices["title"]] if indices["title"] is not None else None) or "",
                    album_title=_safe_str(row_list[indices["album_title"]] if indices["album_title"] is not None else None),
                    barcode=_safe_str(row_list[indices["barcode"]] if indices["barcode"] is not None else None),
                    isrc=_safe_str(row_list[indices["isrc"]] if indices["isrc"] is not None else None),
                    country=_safe_str(row_list[indices["country"]] if indices["country"] is not None else None),
                    shop=_safe_str(row_list[indices["shop"]] if indices["shop"] is not None else None),
                    sales_period=_safe_str(row_list[indices["sales_period"]] if indices["sales_period"] is not None else None),
                    usage_type=_safe_str(row_list[indices["usage_type"]] if indices["usage_type"] is not None else None),
                    sales=_parse_int(row_list[indices["sales"]] if indices["sales"] is not None else 0),
                    returns=_parse_int(row_list[indices["returns"]] if indices["returns"] is not None else 0),
                    ppu=_parse_decimal(row_list[indices["ppu"]] if indices["ppu"] is not None else 0),
                    amount=_parse_decimal(row_list[indices["amount"]] if indices["amount"] is not None else 0),
                    royalty_rate=_parse_percentage(row_list[indices["royalty_rate"]] if indices["royalty_rate"] is not None else None),
                    royalty_amount=_parse_decimal(row_list[indices["royalty_amount"]] if indices["royalty_amount"] is not None else 0),
                    physical_format=None,
                )
                result.rows.append(parsed_row)

            except (ValueError, IndexError, TypeError) as e:
                raw = {}
                for col_name, idx in indices.items():
                    if idx is not None and idx < len(row_list):
                        raw[col_name] = str(row_list[idx]) if row_list[idx] is not None else ""
                result.errors.append(ParseError(
                    row_number=row_number,
                    error=str(e),
                    raw_data=raw,
                ))

    def _parse_physical_sales(self, wb, result: DetailsDetailsParseResult) -> None:
        """Parse the Physical Sales sheet."""
        sheet_name = "Physical Sales"

        # Try to find the sheet (case-insensitive)
        ws = None
        for name in wb.sheetnames:
            if name.lower().strip() == sheet_name.lower():
                ws = wb[name]
                break

        if ws is None:
            # Physical Sales sheet is optional
            return

        rows_iter = ws.iter_rows(values_only=True)

        # Get header row
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return  # Empty sheet, skip

        header_list = list(header_row)
        indices = _find_header_indices(header_list, PHYSICAL_COLUMN_MAPPINGS)

        # Validate required columns
        required = ["artist", "amount"]
        missing = [f for f in required if indices.get(f) is None]
        if missing:
            result.errors.append(ParseError(
                row_number=0,
                error=f"Missing required columns in '{sheet_name}': {missing}. Available: {[str(h) for h in header_list if h]}",
            ))
            return

        for row_number, row in enumerate(rows_iter, start=2):
            result.total_rows += 1
            row_list = list(row)

            # Skip empty rows
            if not row_list or all(cell is None or str(cell).strip() == "" for cell in row_list):
                continue

            try:
                artist = _safe_str(row_list[indices["artist"]] if indices["artist"] is not None else None)
                if not artist:
                    continue  # Skip rows without artist

                # For physical sales, use Identifier as barcode/UPC
                identifier = _safe_str(row_list[indices["identifier"]] if indices["identifier"] is not None else None)
                physical_format = _safe_str(row_list[indices["format"]] if indices["format"] is not None else None)
                usage_type = _safe_str(row_list[indices["type"]] if indices["type"] is not None else None)

                parsed_row = DetailsDetailsRow(
                    row_number=row_number,
                    sheet_name=sheet_name,
                    contract_id=_parse_int(row_list[indices["contract_id"]]) if indices["contract_id"] is not None else None,
                    contract_name=_safe_str(row_list[indices["contract_name"]] if indices["contract_name"] is not None else None),
                    artist=artist,
                    title=_safe_str(row_list[indices["title"]] if indices["title"] is not None else None) or "",
                    album_title=None,  # Not in Physical Sales
                    barcode=identifier,
                    isrc=None,  # Not in Physical Sales
                    country=_safe_str(row_list[indices["country"]] if indices["country"] is not None else None),
                    shop=None,  # Physical Sales doesn't have Shop, use "Physical Sales"
                    sales_period=None,  # Physical Sales doesn't have Sales Period
                    usage_type=usage_type or "Physical",
                    sales=_parse_int(row_list[indices["sales"]] if indices["sales"] is not None else 0),
                    returns=_parse_int(row_list[indices["returns"]] if indices["returns"] is not None else 0),
                    ppu=_parse_decimal(row_list[indices["ppu"]] if indices["ppu"] is not None else 0),
                    amount=_parse_decimal(row_list[indices["amount"]] if indices["amount"] is not None else 0),
                    royalty_rate=_parse_percentage(row_list[indices["royalty_rate"]] if indices["royalty_rate"] is not None else None),
                    royalty_amount=_parse_decimal(row_list[indices["royalty_amount"]] if indices["royalty_amount"] is not None else 0),
                    physical_format=physical_format,
                )
                result.rows.append(parsed_row)

            except (ValueError, IndexError, TypeError) as e:
                raw = {}
                for col_name, idx in indices.items():
                    if idx is not None and idx < len(row_list):
                        raw[col_name] = str(row_list[idx]) if row_list[idx] is not None else ""
                result.errors.append(ParseError(
                    row_number=row_number,
                    error=str(e),
                    raw_data=raw,
                ))
